###############################################################################
### transform_sales.py
### Baca CSV sales data, output Excel:
###   Sheet 1 "Data"    : raw data dari CSV
###   Sheet 2 "Summary" : summary 4 level (AREA / REGION / BRANCH / WOK)
###############################################################################

import os
import sys
import pandas as pd
from datetime import datetime, timedelta
from dotenv import load_dotenv

###############################################################################
# LOAD .env
###############################################################################
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"))

###############################################################################
# CONFIG
###############################################################################
HOMEDIR   = os.getenv("HOMEDIR")
INPUTDIR  = os.path.join(HOMEDIR, "DOWNLOAD", "ps_per_product")
OUTPUTDIR = os.path.join(HOMEDIR, "OUTPUT",   "ps_per_product")

FILE_PREFIX    = "ps_per_product_"
FILE_EXT_IN    = ".csv"
FILE_EXT_OUT   = ".xlsx"
DATE_FORMAT    = "%Y_%m_%d"
RETENTION_DAYS = 7

# Kolom dimensi
DIM_COLS = ["area", "region", "branch", "wok"]

# Kolom metrik dari CSV
METRIC_TODAY    = ["eznet_today", "onedynamic_today", "other_today"]
METRIC_MTD      = ["eznet_mtd", "onedynamic_mtd", "other_mtd"]
METRIC_LAST_MTD = ["eznet_lastmtd", "onedynamic_lastmtd", "other_lastmtd"]
METRIC_TOTAL    = ["total_today", "total_mtd", "total_lastmtd"]

ALL_METRIC_COLS = METRIC_TODAY + METRIC_MTD + METRIC_LAST_MTD + METRIC_TOTAL

# Label kolom output yang rapi
COL_RENAME = {
    "area": "Area",
    "region": "Region",
    "branch": "Branch",
    "wok": "WOK",
    "eznet_today": "EZnet Today",
    "eznet_mtd": "EZnet MTD",
    "eznet_lastmtd": "EZnet Last MTD",
    "onedynamic_today": "One Dynamic Today",
    "onedynamic_mtd": "One Dynamic MTD",
    "onedynamic_lastmtd": "One Dynamic Last MTD",
    "other_today": "Other Today",
    "other_mtd": "Other MTD",
    "other_lastmtd": "Other Last MTD",
    "total_today": "Total Today",
    "total_mtd": "Total MTD",
    "total_lastmtd": "Total Last MTD",
    "contribution_sp": "Contribution SP",
    "mom": "MoM",
}

SHEET_DATA    = "Data"
SHEET_SUMMARY = "Summary"

SUMMARY_LEVELS = [
    {"label": "Summary per AREA",   "dims": ["area"]},
    {"label": "Summary per REGION", "dims": ["area", "region"]},
    {"label": "Summary per BRANCH", "dims": ["area", "region", "branch"]},
    {"label": "Summary per WOK",    "dims": ["area", "region", "branch", "wok"]},
]

###############################################################################
# FUNGSI BANTU
###############################################################################
def cleanup_old_files(folder, prefix, ext, date_fmt, retention_days, log):
    today  = datetime.now().date()
    cutoff = today - timedelta(days=retention_days)
    deleted = 0
    log.info(f"[CLEANUP] Cek file lebih dari {retention_days} hari di: {folder}")
    for fname in os.listdir(folder):
        if not (fname.startswith(prefix) and fname.endswith(ext)):
            continue
        date_str = fname[len(prefix):-len(ext)]
        try:
            file_date = datetime.strptime(date_str, date_fmt).date()
        except ValueError:
            log.warning(f"[CLEANUP] Skip '{fname}': format tanggal tidak dikenali")
            continue
        if file_date <= cutoff:
            try:
                os.remove(os.path.join(folder, fname))
                deleted += 1
                log.info(f"[CLEANUP] Deleted: {fname}")
            except Exception as e:
                log.warning(f"[CLEANUP] Gagal hapus '{fname}': {e}")
    log.info(f"[CLEANUP] Selesai. {deleted} file dihapus.")


def read_csv(path, log):
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip().str.lower()
    # Drop kolom derived dari CSV, akan dihitung ulang
    df = df.drop(columns=["contribution_sp", "mom"], errors="ignore")
    log.info(f"[READ] {len(df)} baris, kolom: {list(df.columns)}")
    return df


def calc_derived(df, grand_mtd):
    df = df.copy()
    total_mtd      = df[METRIC_MTD].sum(axis=1)
    total_last_mtd = df[METRIC_LAST_MTD].sum(axis=1)
    sp_today       = df["eznet_today"] + df["onedynamic_today"]
    df["contribution_sp"] = sp_today.where(df["total_today"] != 0) / df["total_today"].replace(0, float("nan"))
    df["mom"] = (total_mtd - total_last_mtd).where(total_last_mtd != 0) / total_last_mtd.replace(0, float("nan"))
    return df


def build_summary_level(df, dims, grand_mtd, log):
    agg = df.groupby(dims, as_index=False)[ALL_METRIC_COLS].sum()

    agg_mtd      = agg[METRIC_MTD].sum(axis=1)
    agg_last_mtd = agg[METRIC_LAST_MTD].sum(axis=1)

    sp_today_sum   = agg["eznet_today"] + agg["onedynamic_today"]
    total_today    = agg["total_today"]
    agg["contribution_sp"] = sp_today_sum.where(total_today != 0) / total_today.replace(0, float("nan"))
    agg["mom"] = (agg_mtd - agg_last_mtd).where(agg_last_mtd != 0) / agg_last_mtd.replace(0, float("nan"))

    # Baris TOTAL
    total_row = {d: "" for d in dims}
    total_row[dims[0]] = "TOTAL"
    for col in ALL_METRIC_COLS:
        total_row[col] = agg[col].sum()

    total_mtd_val      = agg[METRIC_MTD].sum(axis=1).sum()
    total_last_mtd_val = agg[METRIC_LAST_MTD].sum(axis=1).sum()
    total_sp_today    = agg["eznet_today"].sum() + agg["onedynamic_today"].sum()
    total_today_val   = agg["total_today"].sum()
    total_row["contribution_sp"] = total_sp_today / total_today_val if total_today_val != 0 else float("nan")
    total_row["mom"] = (
        (total_mtd_val - total_last_mtd_val) / total_last_mtd_val
        if total_last_mtd_val != 0 else float("nan")
    )

    agg = pd.concat([agg, pd.DataFrame([total_row])], ignore_index=True)
    log.info(f"[SUMMARY] Level {dims}: {len(agg)-1} baris + 1 total")
    return agg


def build_all_summaries(df, log):
    grand_mtd = df[METRIC_MTD].sum(axis=1).sum()
    log.info(f"[SUMMARY] Grand total MTD: {grand_mtd:,.0f}")

    all_dims = ["area", "region", "branch", "wok"]
    out_cols = all_dims + ALL_METRIC_COLS + ["contribution_sp", "mom"]

    frames = []
    for level in SUMMARY_LEVELS:
        dims  = level["dims"]
        label = level["label"]

        label_row = pd.DataFrame([[label] + [""] * (len(out_cols) - 1)], columns=out_cols)
        agg = build_summary_level(df, dims, grand_mtd, log)

        for d in all_dims:
            if d not in agg.columns:
                agg[d] = ""
        agg = agg[out_cols]

        empty_row = pd.DataFrame([[""] * len(out_cols)], columns=out_cols)
        frames.extend([empty_row, label_row, agg])

    result = pd.concat(frames, ignore_index=True)
    return result, out_cols


###############################################################################
# FORMAT EXCEL
###############################################################################
def format_data_sheet(ws, out_cols):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
    FONT_HEADER = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    FONT_NORMAL = Font(name="Arial", size=10)

    pct_cols = {"Contribution SP", "MoM"}
    num_cols = set(COL_RENAME[c] for c in ALL_METRIC_COLS)

    for i, col in enumerate(out_cols, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font      = FONT_HEADER
        cell.fill      = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = max(len(str(col)) + 4, 14)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = FONT_NORMAL
            col_name = out_cols[cell.column - 1] if cell.column <= len(out_cols) else ""
            if col_name in pct_cols and isinstance(cell.value, float):
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_name in num_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def format_summary_sheet(ws, df_summary, out_cols):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
    FILL_TOTAL   = PatternFill("solid", fgColor="D6DCE4")
    FILL_SECTION = PatternFill("solid", fgColor="BDD7EE")

    FONT_HEADER  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    FONT_TOTAL   = Font(bold=True, name="Arial", size=10)
    FONT_SECTION = Font(bold=True, italic=True, name="Arial", size=10)
    FONT_NORMAL  = Font(name="Arial", size=10)

    pct_cols = {"Contribution SP", "MoM"}
    num_cols = set(COL_RENAME[c] for c in ALL_METRIC_COLS)

    for i, col in enumerate(out_cols, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font      = FONT_HEADER
        cell.fill      = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = max(len(str(col)) + 4, 14)

    for row_idx, row in enumerate(df_summary.itertuples(index=False), start=2):
        row_dict  = dict(zip(out_cols, row))
        first_val = str(row_dict.get(out_cols[0], "")).strip()

        is_total   = first_val == "TOTAL"
        is_empty   = first_val == ""
        is_section = (
            not is_total and not is_empty
            and all(str(row_dict.get(c, "")).strip() == "" for c in out_cols[1:4])
        )

        for col in out_cols:
            cell  = ws.cell(row=row_idx, column=out_cols.index(col) + 1)
            value = row_dict[col]

            if isinstance(value, float) and pd.isna(value):
                cell.value = None
            else:
                cell.value = value if value != "" else None

            if is_section:
                cell.font = FONT_SECTION
                cell.fill = FILL_SECTION
            elif is_total:
                cell.font = FONT_TOTAL
                cell.fill = FILL_TOTAL
            else:
                cell.font = FONT_NORMAL

            if col in pct_cols and isinstance(value, float) and not pd.isna(value):
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col in num_cols and isinstance(value, (int, float)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


###############################################################################
# FUNGSI UTAMA
###############################################################################
def run(datefiltername, log):
    os.makedirs(OUTPUTDIR, exist_ok=True)

    input_file  = os.path.join(INPUTDIR,  f"{FILE_PREFIX}{datefiltername}{FILE_EXT_IN}")
    output_file = os.path.join(OUTPUTDIR, f"{FILE_PREFIX}{datefiltername}{FILE_EXT_OUT}")

    log.info(f"[TRANSFORM] Input  : {input_file}")
    log.info(f"[TRANSFORM] Output : {output_file}")

    if not os.path.exists(input_file):
        raise FileNotFoundError(f"File input tidak ditemukan: {input_file}")

    cleanup_old_files(OUTPUTDIR, FILE_PREFIX, FILE_EXT_OUT, DATE_FORMAT, RETENTION_DAYS, log)

    df = read_csv(input_file, log)

    missing_cols = [c for c in DIM_COLS + ALL_METRIC_COLS if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Kolom tidak ditemukan di CSV: {missing_cols}")

    grand_mtd = df[METRIC_MTD].sum(axis=1).sum()

    # Sheet Data
    df_data       = calc_derived(df, grand_mtd)
    data_raw_cols = DIM_COLS + ALL_METRIC_COLS + ["contribution_sp", "mom"]
    df_data_out   = df_data[data_raw_cols].rename(columns=COL_RENAME)
    data_out_cols = list(df_data_out.columns)

    # Sheet Summary
    log.info("[SUMMARY] Membangun summary semua level ...")
    df_summary_raw, summary_raw_cols = build_all_summaries(df, log)
    df_summary_out   = df_summary_raw.rename(columns=COL_RENAME)
    summary_out_cols = [COL_RENAME.get(c, c) for c in summary_raw_cols]

    # Tulis Excel
    log.info(f"[WRITE] Menulis Excel: {output_file}")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_data_out.to_excel(writer, sheet_name=SHEET_DATA, index=False)
        df_summary_out.to_excel(writer, sheet_name=SHEET_SUMMARY, index=False)

        wb = writer.book
        format_data_sheet(wb[SHEET_DATA], data_out_cols)
        format_summary_sheet(wb[SHEET_SUMMARY], df_summary_out, summary_out_cols)

    log.info(f"[WRITE] Selesai. Output: {output_file}")
    return output_file


###############################################################################
# STANDALONE
###############################################################################
if __name__ == "__main__":
    import logging

    datefiltername = (datetime.now() - timedelta(days=1)).strftime("%Y_%m_%d")

    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[logging.StreamHandler(sys.stdout)],
    )
    log = logging.getLogger()

    try:
        result = run(datefiltername, log)
        log.info(f"[DONE] File output: {result}")
    except Exception as e:
        log.error(f"[FAILED] {e}")
        sys.exit(1)